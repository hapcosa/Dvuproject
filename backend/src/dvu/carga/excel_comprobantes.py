"""El Excel de cobranza, generado desde la base.

Es la misma planilla que hoy produce el bot de WhatsApp a partir del OCR: mismas
columnas, mismos estados, mismos colores. Se conserva la forma a propósito — el área de
cobranza la lee todos los días y aprenderse una planilla nueva no le aporta nada.

Lo que cambia es de dónde salen los datos: ya no de interpretar una foto, sino de lo que
el vendedor declaró en el formulario. Por eso desaparecen las columnas «Texto OCR» y
«Cuenta Origen/Destino», que existían sólo para auditar el OCR.
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from dvu.db.models import Comprobante
from dvu.domain.comprobante import COLORES, ETIQUETAS, EstadoComprobante

FORMATO_CLP = "#,##0"
FORMATO_FECHA = "dd-mm-yyyy"

_ENCABEZADO_FUENTE = Font(bold=True, color="FFFFFF")
_ENCABEZADO_FONDO = PatternFill("solid", fgColor="1E293B")
_BORDE = Border(*(Side(border_style="thin", color="CCCCCC"),) * 4)

COLUMNAS: tuple[tuple[str, int], ...] = (
    ("Fecha aviso", 12),
    ("Vendedor", 22),
    ("Cliente", 38),
    ("RUT cliente", 14),
    ("Factura(s)", 18),
    ("Monto transferido", 16),
    ("Banco", 18),
    ("RUT contraparte", 15),
    ("Fecha transferencia", 16),
    ("N° operación", 16),
    ("Detalle del vendedor", 50),
    ("Comprobante", 30),
    ("Estado", 20),
    ("Observación", 34),
    ("Ingresado", 11),
)


def exportar_comprobantes(
    session: Session,
    *,
    desde: date | None = None,
    hasta: date | None = None,
    incluir_ingresados: bool = True,
) -> bytes:
    """Devuelve el .xlsx en memoria. No escribe archivos: el que descarga decide dónde."""
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None  # noqa: S101 — openpyxl siempre crea la primera hoja
    hoja.title = "Comprobantes"

    filas = _consultar(session, desde=desde, hasta=hasta, incluir_ingresados=incluir_ingresados)
    _escribir(hoja, filas)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _consultar(
    session: Session,
    *,
    desde: date | None,
    hasta: date | None,
    incluir_ingresados: bool,
) -> Sequence[Comprobante]:
    consulta = select(Comprobante).options(
        selectinload(Comprobante.vendedor), selectinload(Comprobante.cliente)
    )
    # Se filtra por la fecha del aviso y no por la de la transferencia: la de
    # transferencia puede venir vacía, y un comprobante sin fecha desaparecería del
    # reporte justo cuando más falta hace mirarlo.
    if desde is not None:
        consulta = consulta.where(Comprobante.creado_en >= desde)
    if hasta is not None:
        consulta = consulta.where(Comprobante.creado_en < _dia_siguiente(hasta))
    if not incluir_ingresados:
        consulta = consulta.where(Comprobante.ingresado.is_(False))
    return session.scalars(consulta.order_by(Comprobante.creado_en.desc())).all()


def _escribir(hoja: Worksheet, filas: Sequence[Comprobante]) -> None:
    for indice, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=indice, value=titulo)
        celda.font = _ENCABEZADO_FUENTE
        celda.fill = _ENCABEZADO_FONDO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = _BORDE
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    hoja.freeze_panes = "A2"
    hoja.row_dimensions[1].height = 28

    for numero, comprobante in enumerate(filas, start=2):
        relleno = _relleno_de(comprobante.estado)
        for indice, valor in enumerate(_valores(comprobante), start=1):
            celda = hoja.cell(row=numero, column=indice, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)
            celda.border = _BORDE
            if relleno is not None:
                celda.fill = relleno
            if indice == 6:  # Monto transferido
                celda.number_format = FORMATO_CLP
            if indice in (1, 9):  # Fechas
                celda.number_format = FORMATO_FECHA

    ultima = get_column_letter(len(COLUMNAS))
    hoja.auto_filter.ref = f"A1:{ultima}{max(1, len(filas) + 1)}"


def _valores(comprobante: Comprobante) -> tuple[str | int | date | None, ...]:
    cliente = comprobante.cliente
    return (
        comprobante.creado_en.date(),
        comprobante.vendedor.nombre,
        cliente.razon_social if cliente is not None else comprobante.cliente_texto,
        cliente.rut if cliente is not None else "",
        ", ".join(comprobante.facturas),
        int(comprobante.monto_clp) if comprobante.monto_clp is not None else None,
        comprobante.banco or "",
        comprobante.rut_contraparte or "",
        comprobante.fecha_transferencia,
        comprobante.numero_operacion or "",
        comprobante.detalle,
        comprobante.imagen_key or "",
        _etiqueta_de(comprobante.estado),
        comprobante.observacion,
        "SÍ" if comprobante.ingresado else "",
    )


def _etiqueta_de(estado: str) -> str:
    """Un estado desconocido se muestra crudo en vez de reventar el reporte: cobranza
    prefiere una planilla rara a ninguna planilla."""
    try:
        return ETIQUETAS[EstadoComprobante(estado)]
    except ValueError:
        return estado


def _relleno_de(estado: str) -> PatternFill | None:
    try:
        return PatternFill("solid", fgColor=COLORES[EstadoComprobante(estado)])
    except ValueError:
        return None


def _dia_siguiente(dia: date) -> date:
    from datetime import timedelta

    return dia + timedelta(days=1)
