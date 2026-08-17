"""Exportación a Excel.

Hoy alguien transcribe ventas y pagos a mano desde WhatsApp. Este módulo genera ese
mismo Excel a partir de los datos ya registrados, para que el dueño siga trabajando
con la planilla que conoce mientras el sistema se asienta.

Tres hojas, en el orden en que se miran: **Ventas** (una fila por pedido), **Detalle**
(una por línea, para revisar qué se vendió) y **Pagos** (con lo aplicado a cada
pedido y el saldo pendiente).

Los montos van como números enteros con formato `#,##0` en CLP: nada de texto, para
que las fórmulas del dueño sigan funcionando. Nunca se escriben decimales.
"""

from __future__ import annotations

from collections.abc import Sequence
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from dvu.db.models import Pago, PagoAplicacion, Pedido
from dvu.domain.pedido import EstadoPedido

FORMATO_CLP = "#,##0"
FORMATO_FECHA = "dd-mm-yyyy"

_ENCABEZADO = Font(bold=True, color="FFFFFF")
_FONDO = PatternFill("solid", fgColor="1F4E79")

COLUMNAS_VENTAS = (
    ("Nº pedido", 16),
    ("Fecha", 12),
    ("RUT cliente", 14),
    ("Cliente", 38),
    ("Vendedor", 22),
    ("Estado", 14),
    ("Origen", 14),
    ("Neto", 14),
    ("IVA", 12),
    ("Total", 14),
    ("Pagado", 14),
    ("Saldo", 14),
)

COLUMNAS_DETALLE = (
    ("Nº pedido", 16),
    ("Fecha", 12),
    ("Cliente", 38),
    ("SKU", 18),
    ("Descripción", 48),
    ("Cantidad", 10),
    ("Múltiplo", 10),
    ("Precio unitario", 16),
    ("Total línea", 14),
)

COLUMNAS_PAGOS = (
    ("Fecha", 12),
    ("RUT cliente", 14),
    ("Cliente", 38),
    ("Método", 16),
    ("Referencia", 22),
    ("Monto", 14),
    ("Estado", 20),
    ("Aplicado a", 30),
    ("Sin aplicar", 14),
)


def exportar_excel(
    session: Session, *, desde: date | None = None, hasta: date | None = None
) -> bytes:
    """Devuelve el .xlsx listo para descargar."""
    libro = Workbook()
    libro.remove(libro.active)  # type: ignore[arg-type]

    pedidos = _pedidos(session, desde, hasta)
    pagos = _pagos(session, desde, hasta)
    pagado_por_pedido = _pagado_por_pedido(pagos)

    _hoja_ventas(libro, pedidos, pagado_por_pedido)
    _hoja_detalle(libro, pedidos)
    _hoja_pagos(libro, pagos)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _pedidos(session: Session, desde: date | None, hasta: date | None) -> list[Pedido]:
    consulta = select(Pedido).options(
        selectinload(Pedido.lineas),
        selectinload(Pedido.cliente),
        selectinload(Pedido.vendedor),
    )
    # Los pedidos anulados quedan fuera del Excel de ventas, pero siguen en la base:
    # la planilla es un reporte, no la fuente de verdad. Los borradores tampoco van:
    # son la lista que el vendedor todavía está armando con el cliente al lado, no
    # tienen folio y nadie los pidió — sumarlos a las ventas sería contar humo.
    consulta = consulta.where(Pedido.estado.notin_(("anulado", EstadoPedido.BORRADOR.value)))
    if desde is not None:
        consulta = consulta.where(Pedido.creado_en >= desde)
    if hasta is not None:
        consulta = consulta.where(Pedido.creado_en <= hasta)
    return list(session.scalars(consulta.order_by(Pedido.id)))


def _pagos(session: Session, desde: date | None, hasta: date | None) -> list[Pago]:
    consulta = select(Pago).options(
        selectinload(Pago.aplicaciones).selectinload(PagoAplicacion.pedido),
        selectinload(Pago.cliente),
    )
    if desde is not None:
        consulta = consulta.where(Pago.fecha_pago >= desde)
    if hasta is not None:
        consulta = consulta.where(Pago.fecha_pago <= hasta)
    return list(session.scalars(consulta.order_by(Pago.fecha_pago, Pago.id)))


def _pagado_por_pedido(pagos: Sequence[Pago]) -> dict[int, int]:
    """Sólo cuenta lo verificado: un pago declarado todavía no es plata en la cuenta."""
    acumulado: dict[int, int] = {}
    for pago in pagos:
        if pago.estado != "verificado":
            continue
        for aplicacion in pago.aplicaciones:
            acumulado[aplicacion.pedido_id] = acumulado.get(aplicacion.pedido_id, 0) + int(
                aplicacion.monto_clp
            )
    return acumulado


def _hoja(libro: Workbook, titulo: str, columnas: Sequence[tuple[str, int]]) -> Worksheet:
    hoja: Worksheet = libro.create_sheet(titulo)
    hoja.append([nombre for nombre, _ in columnas])
    for indice, (_, ancho) in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=indice)
        celda.font = _ENCABEZADO
        celda.fill = _FONDO
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}1"
    return hoja


def _formatear_fila(hoja: Worksheet, fila: int, formatos: dict[int, str]) -> None:
    for columna, formato in formatos.items():
        hoja.cell(row=fila, column=columna).number_format = formato


def _hoja_ventas(libro: Workbook, pedidos: Sequence[Pedido], pagado: dict[int, int]) -> None:
    hoja = _hoja(libro, "Ventas", COLUMNAS_VENTAS)
    formatos = {2: FORMATO_FECHA, **dict.fromkeys((8, 9, 10, 11, 12), FORMATO_CLP)}

    for pedido in pedidos:
        cobrado = pagado.get(pedido.id, 0)
        hoja.append(
            [
                pedido.numero,
                pedido.creado_en.date() if pedido.creado_en else None,
                pedido.cliente.rut,
                pedido.cliente.razon_social,
                pedido.vendedor.nombre if pedido.vendedor else None,
                pedido.estado,
                pedido.origen,
                _entero(pedido.neto_clp),
                _entero(pedido.iva_clp),
                _entero(pedido.total_clp),
                cobrado,
                _entero(pedido.total_clp) - cobrado,
            ]
        )
        _formatear_fila(hoja, hoja.max_row, formatos)

    _totalizar(hoja, columnas=(8, 9, 10, 11, 12), etiqueta_en=7)


def _hoja_detalle(libro: Workbook, pedidos: Sequence[Pedido]) -> None:
    hoja = _hoja(libro, "Detalle", COLUMNAS_DETALLE)
    formatos = {2: FORMATO_FECHA, 8: FORMATO_CLP, 9: FORMATO_CLP}

    for pedido in pedidos:
        for linea in pedido.lineas:
            hoja.append(
                [
                    pedido.numero,
                    pedido.creado_en.date() if pedido.creado_en else None,
                    pedido.cliente.razon_social,
                    linea.sku,
                    linea.descripcion,
                    linea.cantidad,
                    linea.multiplo_venta,
                    _entero(linea.precio_unitario_clp),
                    _entero(linea.total_linea_clp),
                ]
            )
            _formatear_fila(hoja, hoja.max_row, formatos)

    _totalizar(hoja, columnas=(9,), etiqueta_en=8)


def _hoja_pagos(libro: Workbook, pagos: Sequence[Pago]) -> None:
    hoja = _hoja(libro, "Pagos", COLUMNAS_PAGOS)
    formatos = {1: FORMATO_FECHA, 6: FORMATO_CLP, 9: FORMATO_CLP}

    for pago in pagos:
        aplicado = sum(int(a.monto_clp) for a in pago.aplicaciones)
        hoja.append(
            [
                pago.fecha_pago,
                pago.cliente.rut,
                pago.cliente.razon_social,
                pago.metodo,
                pago.referencia,
                _entero(pago.monto_clp),
                pago.estado,
                # Sin folio no hay a qué aplicar un pago, así que la guarda no descarta
                # nada real: es para el tipo, que admite nulo por los borradores.
                ", ".join(a.pedido.numero for a in pago.aplicaciones if a.pedido.numero),
                _entero(pago.monto_clp) - aplicado,
            ]
        )
        _formatear_fila(hoja, hoja.max_row, formatos)

    _totalizar(hoja, columnas=(6, 9), etiqueta_en=5)


def _totalizar(hoja: Worksheet, *, columnas: Sequence[int], etiqueta_en: int) -> None:
    """Suma con fórmula, no con un número calculado aquí: si el dueño filtra o edita
    una fila, el total lo sigue."""
    if hoja.max_row < 2:
        return

    fila = hoja.max_row + 1
    etiqueta = hoja.cell(row=fila, column=etiqueta_en, value="TOTAL")
    etiqueta.font = Font(bold=True)

    for columna in columnas:
        letra = get_column_letter(columna)
        celda = hoja.cell(row=fila, column=columna, value=f"=SUM({letra}2:{letra}{fila - 1})")
        celda.font = Font(bold=True)
        celda.number_format = FORMATO_CLP


def _entero(valor: Decimal | int | Any) -> int:
    """CLP sin decimales, en toda la planilla."""
    return int(valor or 0)
