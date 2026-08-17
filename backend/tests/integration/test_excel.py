"""El Excel que hoy se llena a mano.

Lo que se verifica es lo que el dueño mira: montos enteros en CLP, el saldo por
pedido y que un pago sólo declarado no cuente como cobrado.
"""

from __future__ import annotations

import uuid as uuid_lib
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from dvu.carga.excel import exportar_excel
from dvu.db.models import (
    Cliente,
    Pago,
    PagoAplicacion,
    Pedido,
    PedidoLinea,
    Producto,
    Usuario,
)
from dvu.seguridad import emitir_token, hashear

pytestmark = pytest.mark.integration

PREFIJO = "/api/v1"


@pytest.fixture
def venta(sesion: Session) -> dict[str, Any]:
    """Un pedido de 51.122 CLP con un pago verificado de 30.000 y otro sólo declarado."""
    vendedor = Usuario(
        email="v@test.cl", nombre="Vendedor Uno", rol="vendedor", password_hash=hashear("x")
    )
    admin = Usuario(email="a@test.cl", nombre="Admin", rol="admin", password_hash=hashear("x"))
    sesion.add_all([vendedor, admin])
    sesion.flush()

    cliente = Cliente(rut="76123456-0", razon_social="FERRETERIA TEST SPA", vendedor_id=vendedor.id)
    producto = Producto(
        sku="DVU-PR49573",
        descripcion="LIQUIDO DE FRENO FEDERAL",
        multiplo_venta=12,
        precio_lista_clp=Decimal("1790"),
    )
    sesion.add_all([cliente, producto])
    sesion.flush()

    pedido = Pedido(
        client_uuid=uuid_lib.uuid4(),
        numero="P-2026-000001",
        cliente_id=cliente.id,
        vendedor_id=vendedor.id,
        origen="app_vendedor",
        estado="confirmado",
        neto_clp=Decimal("42960"),
        iva_clp=Decimal("8162"),
        total_clp=Decimal("51122"),
        lineas=[
            PedidoLinea(
                producto_id=producto.id,
                sku=producto.sku,
                descripcion=producto.descripcion,
                multiplo_venta=12,
                cantidad=24,
                precio_unitario_clp=Decimal("1790"),
                total_linea_clp=Decimal("42960"),
            )
        ],
    )
    sesion.add(pedido)
    sesion.flush()

    verificado = Pago(
        cliente_id=cliente.id,
        monto_clp=Decimal("30000"),
        fecha_pago=date(2026, 8, 1),
        metodo="transferencia",
        referencia="OP-1",
        estado="verificado",
        registrado_por=vendedor.id,
        aplicaciones=[PagoAplicacion(pedido_id=pedido.id, monto_clp=Decimal("30000"))],
    )
    declarado = Pago(
        cliente_id=cliente.id,
        monto_clp=Decimal("21122"),
        fecha_pago=date(2026, 8, 2),
        metodo="transferencia",
        referencia="OP-2",
        estado="declarado",
        registrado_por=vendedor.id,
        aplicaciones=[PagoAplicacion(pedido_id=pedido.id, monto_clp=Decimal("21122"))],
    )
    sesion.add_all([verificado, declarado])
    sesion.flush()

    return {
        "cliente": cliente,
        "pedido": pedido,
        "vendedor": vendedor,
        "auth_admin": {"Authorization": f"Bearer {emitir_token(admin.uuid, 'admin')}"},
        "auth_vendedor": {"Authorization": f"Bearer {emitir_token(vendedor.uuid, 'vendedor')}"},
    }


def _abrir(contenido: bytes) -> Any:
    return load_workbook(BytesIO(contenido))


def test_tres_hojas_en_el_orden_en_que_se_miran(sesion: Session, venta: dict[str, Any]) -> None:
    libro = _abrir(exportar_excel(sesion))
    assert libro.sheetnames == ["Ventas", "Detalle", "Pagos"]


def test_ventas_trae_el_pedido_con_montos_enteros(sesion: Session, venta: dict[str, Any]) -> None:
    hoja = _abrir(exportar_excel(sesion))["Ventas"]
    fila = [c.value for c in hoja[2]]

    assert fila[0] == "P-2026-000001"
    assert fila[2] == "76123456-0"
    assert fila[4] == "Vendedor Uno"
    # Neto, IVA y total: enteros, nunca float ni texto.
    assert (fila[7], fila[8], fila[9]) == (42960, 8162, 51122)
    assert all(isinstance(v, int) for v in fila[7:12])


def test_solo_lo_verificado_cuenta_como_pagado(sesion: Session, venta: dict[str, Any]) -> None:
    """El pago declarado todavía no es plata en la cuenta: no baja el saldo."""
    hoja = _abrir(exportar_excel(sesion))["Ventas"]
    fila = [c.value for c in hoja[2]]

    assert fila[10] == 30000  # Pagado
    assert fila[11] == 51122 - 30000  # Saldo


def test_los_totales_son_formulas(sesion: Session, venta: dict[str, Any]) -> None:
    """Si el dueño filtra o edita una fila, el total tiene que seguirlo."""
    hoja = _abrir(exportar_excel(sesion))["Ventas"]
    fila_total = hoja.max_row

    assert hoja.cell(row=fila_total, column=7).value == "TOTAL"
    assert hoja.cell(row=fila_total, column=10).value == "=SUM(J2:J2)"


def test_detalle_conserva_el_multiplo_de_venta(sesion: Session, venta: dict[str, Any]) -> None:
    hoja = _abrir(exportar_excel(sesion))["Detalle"]
    fila = [c.value for c in hoja[2]]

    assert fila[3] == "DVU-PR49573"
    assert (fila[5], fila[6]) == (24, 12)  # cantidad y múltiplo
    assert fila[8] == 42960


def test_pagos_muestra_a_que_pedido_se_aplico(sesion: Session, venta: dict[str, Any]) -> None:
    hoja = _abrir(exportar_excel(sesion))["Pagos"]
    filas = [[c.value for c in f] for f in hoja.iter_rows(min_row=2, max_row=3)]

    assert [f[6] for f in filas] == ["verificado", "declarado"]
    assert all(f[7] == "P-2026-000001" for f in filas)
    assert [f[8] for f in filas] == [0, 0]  # nada sin aplicar


def test_pedido_anulado_no_aparece_pero_sigue_en_la_base(
    sesion: Session, venta: dict[str, Any]
) -> None:
    venta["pedido"].estado = "anulado"
    sesion.flush()

    hoja = _abrir(exportar_excel(sesion))["Ventas"]
    assert hoja.max_row == 1  # sólo el encabezado, sin fila de totales

    assert sesion.get(Pedido, venta["pedido"].id) is not None


def test_una_lista_a_medias_no_es_una_venta(sesion: Session, venta: dict[str, Any]) -> None:
    """El borrador del vendedor no tiene folio y nadie lo pidió: sumarlo sería contar humo."""
    venta["pedido"].estado = "borrador"
    venta["pedido"].numero = None
    sesion.flush()

    hoja = _abrir(exportar_excel(sesion))["Ventas"]
    assert hoja.max_row == 1


def test_el_rango_de_fechas_filtra_los_pagos(sesion: Session, venta: dict[str, Any]) -> None:
    hoja = _abrir(exportar_excel(sesion, desde=date(2026, 8, 2)))["Pagos"]
    filas = [[c.value for c in f] for f in hoja.iter_rows(min_row=2, max_row=2)]

    assert len(filas) == 1
    assert filas[0][4] == "OP-2"


def test_descarga_del_xlsx_solo_para_admin(cliente_api: TestClient, venta: dict[str, Any]) -> None:
    vendedor = cliente_api.get(f"{PREFIJO}/reportes/ventas.xlsx", headers=venta["auth_vendedor"])
    assert vendedor.status_code == 403

    admin = cliente_api.get(f"{PREFIJO}/reportes/ventas.xlsx", headers=venta["auth_admin"])
    assert admin.status_code == 200
    assert admin.headers["content-disposition"].startswith("attachment;")
    assert _abrir(admin.content).sheetnames == ["Ventas", "Detalle", "Pagos"]
