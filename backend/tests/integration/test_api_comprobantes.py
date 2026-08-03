"""Comprobantes de transferencia: alta del vendedor, bandeja de cobranza y Excel.

Es el reemplazo del grupo de WhatsApp. Lo que se prueba acá es sobre todo lo que el
sistema **no** hace: no rechaza un aviso incompleto, no borra nada y no deja que un
vendedor vea los comprobantes de otro.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from dvu.almacenamiento import AlmacenLocal, get_almacen
from dvu.api.main import create_app
from dvu.db.models import Cliente, Comprobante, Usuario
from dvu.db.session import get_session
from dvu.seguridad import emitir_token, hashear

pytestmark = pytest.mark.integration

PREFIJO = "/api/v1"
RUT = "76123456-0"


@pytest.fixture
def datos(sesion: Session) -> dict[str, Any]:
    vendedor = Usuario(
        email="v@test.cl", nombre="Vendedor", rol="vendedor", password_hash=hashear("x")
    )
    otro = Usuario(email="v2@test.cl", nombre="Otro", rol="vendedor", password_hash=hashear("x"))
    admin = Usuario(email="a@test.cl", nombre="Cobranza", rol="admin", password_hash=hashear("x"))
    sesion.add_all([vendedor, otro, admin])
    sesion.flush()

    cliente = Cliente(rut=RUT, razon_social="FERRETERIA EL MARTILLO SPA", vendedor_id=vendedor.id)
    sesion.add(cliente)
    sesion.flush()

    return {
        "vendedor": vendedor,
        "otro": otro,
        "admin": admin,
        "cliente": cliente,
        "auth_vendedor": {"Authorization": f"Bearer {emitir_token(vendedor.uuid, 'vendedor')}"},
        "auth_otro": {"Authorization": f"Bearer {emitir_token(otro.uuid, 'vendedor')}"},
        "auth_admin": {"Authorization": f"Bearer {emitir_token(admin.uuid, 'admin')}"},
    }


@pytest.fixture
def cliente_api_local(sesion: Session, tmp_path: Any) -> Any:
    """Cliente HTTP con almacén en disco: los tests no necesitan MinIO."""
    app = create_app()
    app.dependency_overrides[get_session] = lambda: sesion
    app.dependency_overrides[get_almacen] = lambda: AlmacenLocal(tmp_path)
    with TestClient(app) as cliente:
        yield cliente
    app.dependency_overrides.clear()


def _payload(**extra: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "cliente_texto": "FERRETERIA EL MARTILLO",
        "cliente_rut": RUT,
        "facturas": ["33780"],
        "monto_clp": 510_459,
        "banco": "BCI",
        "numero_operacion": "12345678",
        "fecha_transferencia": str(date(2026, 8, 1)),
        "detalle": "pago factura 33780",
    }
    base.update(extra)
    return base


# --- alta ---------------------------------------------------------------------


def test_comprobante_completo_queda_listo(cliente_api: TestClient, datos: dict[str, Any]) -> None:
    r = cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])

    assert r.status_code == 201
    cuerpo = r.json()
    assert cuerpo["estado"] == "listo"
    assert cuerpo["estado_etiqueta"] == "LISTO PARA INGRESAR"
    # El RUT calzó: el comprobante queda amarrado al cliente de la base.
    assert cuerpo["cliente"] == "FERRETERIA EL MARTILLO SPA"
    assert cuerpo["cliente_rut"] == RUT


def test_comprobante_incompleto_se_guarda_marcado(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    """Perder el aviso es peor que registrarlo a medias: nunca devuelve un error."""
    r = cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json=_payload(monto_clp=None, detalle="transferí lo de siempre"),
        headers=datos["auth_vendedor"],
    )

    assert r.status_code == 201
    assert r.json()["estado"] == "falta_monto"


def test_rut_invalido_no_pierde_el_comprobante(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    """Un dígito verificador mal tipeado deja al cliente sin identificar, no descarta
    el aviso completo."""
    r = cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json=_payload(cliente_rut="76123456-9"),
        headers=datos["auth_vendedor"],
    )

    assert r.status_code == 201
    assert r.json()["cliente_rut"] is None
    # Queda el texto que escribió el vendedor: es la evidencia para corregirlo después.
    assert r.json()["cliente"] == "FERRETERIA EL MARTILLO"


def test_monto_y_factura_se_leen_del_texto_libre(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    """En terreno se escribe todo de corrido. Si el dato está en el texto, está."""
    r = cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json={"cliente_texto": "EL MARTILLO", "detalle": "factura 33780 por $510.459"},
        headers=datos["auth_vendedor"],
    )

    assert r.status_code == 201
    cuerpo = r.json()
    assert cuerpo["monto_clp"] == 510_459
    assert cuerpo["facturas"] == ["33780"]
    assert cuerpo["estado"] == "listo"


def test_abono_parcial_se_marca_aunque_este_completo(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    r = cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json=_payload(detalle="abono a cuenta de la 33780"),
        headers=datos["auth_vendedor"],
    )

    assert r.json()["estado"] == "abono_parcial"


def test_reenvio_con_el_mismo_client_uuid_no_duplica(
    cliente_api: TestClient, datos: dict[str, Any], sesion: Session
) -> None:
    """La app del vendedor reintenta al recuperar señal sin saber si el primero llegó."""
    payload = _payload(client_uuid="11111111-1111-1111-1111-111111111111")

    primera = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=payload, headers=datos["auth_vendedor"]
    )
    segunda = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=payload, headers=datos["auth_vendedor"]
    )

    assert primera.json()["uuid"] == segunda.json()["uuid"]
    assert len(sesion.scalars(select(Comprobante)).all()) == 1


def test_misma_operacion_dos_veces_queda_como_posible_duplicado(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])
    segunda = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    )

    assert segunda.json()["estado"] == "duplicado_posible"
    assert "duplicado" in segunda.json()["observacion"].lower()


def test_sin_numero_de_operacion_no_se_marca_duplicado(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    """Dos ferreterías pueden transferir el mismo monto el mismo día. Una alarma que
    suena todos los días deja de mirarse."""
    sin_operacion = _payload(numero_operacion=None)

    cliente_api.post(f"{PREFIJO}/comprobantes", json=sin_operacion, headers=datos["auth_vendedor"])
    segunda = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=sin_operacion, headers=datos["auth_vendedor"]
    )

    assert segunda.json()["estado"] == "listo"


def test_el_cliente_no_puede_registrar_comprobantes(
    cliente_api: TestClient, sesion: Session
) -> None:
    usuario = Usuario(email="c@test.cl", nombre="C", rol="cliente", password_hash=hashear("x"))
    sesion.add(usuario)
    sesion.flush()
    auth = {"Authorization": f"Bearer {emitir_token(usuario.uuid, 'cliente')}"}

    r = cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=auth)

    assert r.status_code == 403


# --- bandeja ------------------------------------------------------------------


def test_el_vendedor_solo_ve_los_suyos(cliente_api: TestClient, datos: dict[str, Any]) -> None:
    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])
    cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json=_payload(numero_operacion="99999999"),
        headers=datos["auth_otro"],
    )

    propios = cliente_api.get(f"{PREFIJO}/comprobantes", headers=datos["auth_vendedor"]).json()
    todos = cliente_api.get(f"{PREFIJO}/comprobantes", headers=datos["auth_admin"]).json()

    assert propios["total"] == 1
    assert propios["items"][0]["vendedor"] == "Vendedor"
    assert todos["total"] == 2


def test_filtro_por_estado_y_pendientes(cliente_api: TestClient, datos: dict[str, Any]) -> None:
    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])
    cliente_api.post(
        f"{PREFIJO}/comprobantes",
        json=_payload(monto_clp=None, numero_operacion=None),
        headers=datos["auth_vendedor"],
    )

    r = cliente_api.get(f"{PREFIJO}/comprobantes?estado=falta_monto", headers=datos["auth_admin"])

    assert r.json()["total"] == 1
    assert r.json()["items"][0]["estado"] == "falta_monto"


def test_marcar_ingresado_no_borra_la_fila(
    cliente_api: TestClient, datos: dict[str, Any], sesion: Session
) -> None:
    creado = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    ).json()

    r = cliente_api.post(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/ingresado", headers=datos["auth_admin"]
    )

    assert r.status_code == 200
    assert r.json()["ingresado"] is True
    # Sale de la bandeja pero sigue existiendo.
    assert len(sesion.scalars(select(Comprobante)).all()) == 1
    pendientes = cliente_api.get(
        f"{PREFIJO}/comprobantes?pendientes=true", headers=datos["auth_admin"]
    )
    assert pendientes.json()["total"] == 0


def test_el_vendedor_no_puede_marcar_ingresado(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    creado = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    ).json()

    r = cliente_api.post(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/ingresado", headers=datos["auth_vendedor"]
    )

    assert r.status_code == 403


# --- imagen -------------------------------------------------------------------


def test_subir_y_ver_la_imagen(cliente_api_local: TestClient, datos: dict[str, Any]) -> None:
    creado = cliente_api_local.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    ).json()

    subida = cliente_api_local.post(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/imagen",
        files={"archivo": ("comprobante.png", io.BytesIO(b"\x89PNG imagen"), "image/png")},
        headers=datos["auth_vendedor"],
    )

    assert subida.status_code == 200
    assert subida.json()["tiene_imagen"] is True

    vista = cliente_api_local.get(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/imagen",
        headers=datos["auth_vendedor"],
        follow_redirects=False,
    )
    # Nunca se sirve el contenido directo: la imagen trae datos bancarios del cliente.
    assert vista.status_code == 307


def test_no_se_puede_adjuntar_al_comprobante_de_otro(
    cliente_api_local: TestClient, datos: dict[str, Any]
) -> None:
    creado = cliente_api_local.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    ).json()

    r = cliente_api_local.post(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/imagen",
        files={"archivo": ("x.png", io.BytesIO(b"\x89PNG"), "image/png")},
        headers=datos["auth_otro"],
    )

    assert r.status_code == 403


def test_ver_imagen_sin_token_no_funciona(
    cliente_api_local: TestClient, datos: dict[str, Any]
) -> None:
    creado = cliente_api_local.post(
        f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"]
    ).json()

    r = cliente_api_local.get(
        f"{PREFIJO}/comprobantes/{creado['uuid']}/imagen", follow_redirects=False
    )

    assert r.status_code in (401, 403)


# --- Excel --------------------------------------------------------------------


def test_reporte_xlsx_trae_lo_declarado(cliente_api: TestClient, datos: dict[str, Any]) -> None:
    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])

    r = cliente_api.get(f"{PREFIJO}/comprobantes/reporte.xlsx", headers=datos["auth_admin"])

    assert r.status_code == 200
    assert "attachment" in r.headers["content-disposition"]

    hoja = load_workbook(io.BytesIO(r.content)).active
    assert hoja is not None
    encabezados = [celda.value for celda in hoja[1]]
    assert "Vendedor" in encabezados

    fila = [celda.value for celda in hoja[2]]
    # El monto viaja como número, no como texto: cobranza lo suma en la planilla.
    assert 510_459 in fila
    assert "LISTO PARA INGRESAR" in fila


def test_reporte_xlsx_es_solo_de_administracion(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    r = cliente_api.get(f"{PREFIJO}/comprobantes/reporte.xlsx", headers=datos["auth_vendedor"])

    assert r.status_code == 403


def test_reporte_xlsx_vacio_sigue_siendo_un_excel_valido(
    cliente_api: TestClient, datos: dict[str, Any]
) -> None:
    """Cobranza prefiere una planilla vacía a un error: le dice que no hubo avisos."""
    r = cliente_api.get(f"{PREFIJO}/comprobantes/reporte.xlsx", headers=datos["auth_admin"])

    hoja = load_workbook(io.BytesIO(r.content)).active
    assert hoja is not None
    assert hoja.max_row == 1


def test_el_comprobante_no_crea_un_pago(
    cliente_api: TestClient, datos: dict[str, Any], sesion: Session
) -> None:
    """El comprobante es lo que declaró el vendedor; el pago es el hecho contable. Que
    exista uno no puede fabricar el otro sin que nadie lo verifique."""
    from dvu.db.models import Pago

    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])

    assert sesion.scalars(select(Pago)).all() == []


def test_monto_negativo_o_cero_se_rechaza(cliente_api: TestClient, datos: dict[str, Any]) -> None:
    """Lo único que sí se rechaza: un monto imposible no es un dato incompleto, es un
    dato erróneo, y guardarlo ensucia el total que cobranza suma."""
    r = cliente_api.post(
        f"{PREFIJO}/comprobantes", json=_payload(monto_clp=0), headers=datos["auth_vendedor"]
    )

    assert r.status_code == 422


def test_el_monto_se_guarda_entero_sin_decimales(
    cliente_api: TestClient, datos: dict[str, Any], sesion: Session
) -> None:
    cliente_api.post(f"{PREFIJO}/comprobantes", json=_payload(), headers=datos["auth_vendedor"])

    comprobante = sesion.scalars(select(Comprobante)).one()
    assert comprobante.monto_clp == Decimal("510459")
    assert comprobante.monto_clp == comprobante.monto_clp.to_integral_value()
